import json
import base64
import xmlsec
from pathlib import Path
from lxml import etree
import xml.etree.ElementTree as ET
from xml.dom import minidom
from tkinter import filedialog, messagebox
from cryptography.hazmat.primitives.asymmetric import dsa
from cryptography.hazmat.primitives import serialization
import openpyxl
from openpyxl.cell import MergedCell
import customtkinter as ctk

ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")


class ExcelToXMLUltimate(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("Excel → XML + Digital Signature (CoA Ready)")
        self.geometry("1450x950")
        self.excel_path = None
        self.elements = []
        self.private_key_pem = None
        self.public_key_pem = None
        self.xmlsec_key = None  # Store xmlsec key object
        self.setup_ui()

    def setup_ui(self):
        ctk.CTkLabel(self, text="Excel to Signed XML Generator", font=("Helvetica", 28, "bold")).pack(pady=20)

        file_frame = ctk.CTkFrame(self)
        file_frame.pack(fill="x", padx=30, pady=(0, 10))
        self.excel_label = ctk.CTkLabel(file_frame, text="No Excel file selected", font=("Arial", 14), anchor="w")
        self.excel_label.pack(side="left", fill="x", expand=True, padx=20)
        ctk.CTkButton(file_frame, text="Browse Excel", width=160, command=self.browse_excel).pack(side="right", padx=20)

        self.container = ctk.CTkScrollableFrame(self)
        self.container.pack(fill="both", expand=True, padx=30, pady=10)

        btn_frame = ctk.CTkFrame(self)
        btn_frame.pack(fill="x", padx=30, pady=15)

        ctk.CTkButton(btn_frame, text="+ Single Field", fg_color="#1f6aa5", height=40, command=self.add_single).pack(side="left", padx=6)
        ctk.CTkButton(btn_frame, text="+ Repeated Block", fg_color="#883997", height=40, command=self.add_repeated).pack(side="left", padx=6)
        ctk.CTkButton(btn_frame, text="+ Nested Block", fg_color="#2b8a3e", height=40, command=self.add_nested).pack(side="left", padx=6)
        ctk.CTkButton(btn_frame, text="Save Template", command=self.save_config).pack(side="left", padx=6)
        ctk.CTkButton(btn_frame, text="Load Template", command=self.load_config).pack(side="left", padx=6)
        ctk.CTkButton(btn_frame, text="Sign XML", fg_color="green", command=self.sign_xml_with_dialog).pack(side="left", padx=6)
        ctk.CTkButton(btn_frame, text="Verify XML", fg_color="orange", command=self.verify_xml).pack(side="left", padx=6)
        ctk.CTkButton(btn_frame, text="GENERATE XML", width=220, height=50, fg_color="darkorange",
                      font=("Helvetica", 18, "bold"), command=self.generate_xml).pack(side="right", padx=20)

    def remove_element(self, frame):
        for elem in self.elements[:]:
            if elem["frame"] == frame:
                frame.destroy()
                self.elements.remove(elem)
                break

    def add_single(self):
        frame = ctk.CTkFrame(self.container)
        frame.pack(fill="x", pady=6, padx=10)
        ctk.CTkLabel(frame, text="Single", width=90, fg_color="#1f6aa5", text_color="white", corner_radius=8).pack(side="left", padx=10)
        node = ctk.CTkEntry(frame, placeholder_text="Tag name", width=220)
        row = ctk.CTkEntry(frame, placeholder_text="Row", width=80)
        col = ctk.CTkEntry(frame, placeholder_text="Col", width=80)
        fixed = ctk.CTkEntry(frame, placeholder_text="Fixed value (optional)", width=350)
        node.pack(side="left", padx=5); row.pack(side="left", padx=5); col.pack(side="left", padx=5); fixed.pack(side="left", padx=5, fill="x", expand=True)
        ctk.CTkButton(frame, text="X", width=40, fg_color="red", command=lambda f=frame: self.remove_element(f)).pack(side="right", padx=10)
        self.elements.append({"type": "single", "frame": frame, "node": node, "row": row, "col": col, "fixed": fixed})

    def add_repeated(self):
        frame = ctk.CTkFrame(self.container, fg_color="#f8f9fa")
        frame.pack(fill="x", pady=12, padx=10)
        header = ctk.CTkFrame(frame)
        header.pack(fill="x", pady=5)
        ctk.CTkLabel(header, text="Repeated", fg_color="#883997", text_color="white", corner_radius=8).pack(side="left", padx=10)
        name = ctk.CTkEntry(header, placeholder_text="Block name (e.g. BatchAnalysis)", width=280)
        name.pack(side="left", padx=10)
        ctk.CTkLabel(header, text="Rows:").pack(side="left")
        start = ctk.CTkEntry(header, width=80); start.insert(0, "2")
        start.pack(side="left", padx=5)
        ctk.CTkLabel(header, text="to:").pack(side="left")
        end = ctk.CTkEntry(header, width=80, placeholder_text="blank=auto")
        end.pack(side="left", padx=5)
        ctk.CTkButton(header, text="Remove", fg_color="red", command=lambda f=frame: self.remove_element(f)).pack(side="right", padx=10)

        fields_frame = ctk.CTkFrame(frame)
        fields_frame.pack(fill="x", padx=40, pady=5)
        fields = []

        def add_field():
            f = ctk.CTkFrame(fields_frame)
            f.pack(fill="x", pady=3)
            node_e = ctk.CTkEntry(f, placeholder_text="Field", width=220)
            col_e = ctk.CTkEntry(f, placeholder_text="Col", width=80)
            off_e = ctk.CTkEntry(f, placeholder_text="Offset", width=100); off_e.insert(0, "0")
            node_e.pack(side="left", padx=5); col_e.pack(side="left", padx=5); off_e.pack(side="left", padx=5)
            
            def remove_field(field_dict, field_frame):
                field_frame.destroy()
                if field_dict in fields:
                    fields.remove(field_dict)
            
            field_dict = {"node": node_e, "col": col_e, "offset": off_e}
            ctk.CTkButton(f, text="X", width=40, fg_color="red", 
                         command=lambda: remove_field(field_dict, f)).pack(side="right", padx=5)
            fields.append(field_dict)

        ctk.CTkButton(frame, text="+ Add Field", fg_color="green", command=add_field).pack(pady=8)
        add_field()
        self.elements.append({"type": "repeated", "frame": frame, "name": name, "start": start, 
                            "end": end, "fields": fields, "fields_frame": fields_frame})

    def add_nested(self):
        frame = ctk.CTkFrame(self.container, fg_color="#e8f5e9")
        frame.pack(fill="x", pady=12, padx=10)
        header = ctk.CTkFrame(frame)
        header.pack(fill="x", pady=5)
        ctk.CTkLabel(header, text="Nested", fg_color="#2b8a3e", text_color="white", corner_radius=8).pack(side="left", padx=10)
        block_name = ctk.CTkEntry(header, placeholder_text="Block name (e.g. FileSignature)", width=320)
        block_name.pack(side="left", padx=10)
        ctk.CTkButton(header, text="Remove", fg_color="red", command=lambda f=frame: self.remove_element(f)).pack(side="right", padx=10)

        subs_frame = ctk.CTkFrame(frame)
        subs_frame.pack(fill="x", padx=50, pady=5)
        subs = []

        def add_sub():
            f = ctk.CTkFrame(subs_frame)
            f.pack(fill="x", pady=3)
            tag = ctk.CTkEntry(f, placeholder_text="Tag", width=250)
            val = ctk.CTkEntry(f, placeholder_text="Fixed value", width=200)
            row = ctk.CTkEntry(f, placeholder_text="Row", width=80)
            col = ctk.CTkEntry(f, placeholder_text="Col", width=80)
            tag.pack(side="left", padx=5); val.pack(side="left", padx=5); row.pack(side="left", padx=5); col.pack(side="left", padx=5)
            
            def remove_sub(sub_dict, sub_frame):
                sub_frame.destroy()
                if sub_dict in subs:
                    subs.remove(sub_dict)
            
            sub_dict = {"tag": tag, "value": val, "row": row, "col": col}
            ctk.CTkButton(f, text="X", width=40, fg_color="red", 
                         command=lambda: remove_sub(sub_dict, f)).pack(side="right", padx=5)
            subs.append(sub_dict)

        ctk.CTkButton(frame, text="+ Add Sub-tag", fg_color="green", command=add_sub).pack(pady=8)
        add_sub()
        self.elements.append({"type": "nested", "frame": frame, "block_name": block_name, 
                            "subnodes": subs, "subs_frame": subs_frame})

    def browse_excel(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx *.xls")])
        if path:
            self.excel_path = path
            self.excel_label.configure(text=f"Excel: {Path(path).name}")

    def get_cell(self, sheet, r, c):
        try:
            cell = sheet.cell(row=int(r), column=int(c))
            if isinstance(cell, MergedCell):
                for mr in sheet.merged_cells.ranges:
                    if cell.coordinate in mr:
                        cell = sheet.cell(mr.min_row, mr.min_col)
                        break
            v = cell.value
            return "" if v is None else str(v).strip()
        except:
            return ""

    def _safe_get(self, widget):
        try:
            if widget.winfo_exists():
                return widget.get()
        except:
            pass
        return ""

    def generate_xml(self):
        if not self.excel_path:
            messagebox.showerror("Error", "Select Excel file first!")
            return None

        wb = None
        try:
            wb = openpyxl.load_workbook(self.excel_path, read_only=True)
            sheet = wb.active
            root = ET.Element("FileInformation")

            for elem in self.elements:
                if not elem["frame"].winfo_exists():
                    continue

                t = elem["type"]
                if t == "single":
                    tag = self._safe_get(elem["node"]).strip()
                    if not tag: continue
                    text = self._safe_get(elem["fixed"]) or self.get_cell(sheet, self._safe_get(elem["row"]), self._safe_get(elem["col"]))
                    ET.SubElement(root, tag).text = text

                elif t == "repeated":
                    name = self._safe_get(elem["name"]).strip()
                    if not name: continue
                    try:
                        start = int(self._safe_get(elem["start"]) or 2)
                        end_val = self._safe_get(elem["end"])
                        end = int(end_val) if end_val else 99999
                    except:
                        start, end = 2, 99999

                    field_defs = []
                    for f in elem["fields"]:
                        node_name = self._safe_get(f["node"]).strip()
                        col = self._safe_get(f["col"])
                        off = self._safe_get(f["offset"]) or "0"
                        if node_name and col.isdigit():
                            try:
                                field_defs.append((node_name, int(col), int(off)))
                            except:
                                continue

                    for r in range(start, end + 1):
                        if r > start and not sheet.cell(row=r, column=1).value:
                            break
                        block = ET.SubElement(root, name)
                        for tag_name, col, offset in field_defs:
                            val = self.get_cell(sheet, r + offset, col)
                            ET.SubElement(block, tag_name).text = val

                elif t == "nested":
                    block_name = self._safe_get(elem["block_name"]).strip()
                    if not block_name: continue
                    block = ET.SubElement(root, block_name)
                    for s in elem["subnodes"]:
                        tag = self._safe_get(s["tag"]).strip()
                        if not tag: continue
                        fixed = self._safe_get(s["value"])
                        text = fixed or self.get_cell(sheet, self._safe_get(s["row"]), self._safe_get(s["col"]))
                        ET.SubElement(block, tag).text = text

            rough = ET.tostring(root, encoding='utf-8', method='xml')
            dom = minidom.parseString(rough)
            pretty = dom.toprettyxml(indent="  ")
            lines = pretty.splitlines()
            if lines[0].startswith("<?xml"):
                lines.pop(0)
            pretty_no_header = "\n".join(lines).strip() + "\n"
            header = '<?xml version="1.0" encoding="UTF-8"?>\n'
            final_xml = header + pretty_no_header

            save_path = filedialog.asksaveasfilename(defaultextension=".xml", filetypes=[("XML", "*.xml")])
            if save_path:
                with open(save_path, "w", encoding="utf-8") as f:
                    f.write(final_xml)
                messagebox.showinfo("Success", f"XML saved!\n{save_path}")
                return save_path
        except Exception as e:
            messagebox.showerror("Error", f"Failed to generate XML:\n{str(e)}")
            return None
        finally:
            if wb:
                wb.close()

    def generate_keys_internal(self):
        """Internal method to generate DSA key pair and store xmlsec key object"""
        try:
            # Generate DSA key pair using 1024-bit (as per reference)
            private_key = dsa.generate_private_key(key_size=1024)
            
            # Use TraditionalOpenSSL format (as per reference)
            self.private_key_pem = private_key.private_bytes(
                encoding=serialization.Encoding.PEM,
                format=serialization.PrivateFormat.TraditionalOpenSSL,
                encryption_algorithm=serialization.NoEncryption()
            )
            
            public_key = private_key.public_key()
            self.public_key_pem = public_key.public_bytes(
                encoding=serialization.Encoding.PEM,
                format=serialization.PublicFormat.SubjectPublicKeyInfo
            )
            
            # Load key into xmlsec Key object and store it
            self.xmlsec_key = xmlsec.Key.from_memory(
                self.private_key_pem, 
                xmlsec.constants.KeyDataFormatPem, 
                None
            )
            
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            raise Exception(f"Failed to generate keys: {str(e)}\n\n{error_details}")

    def sign_xml_with_dialog(self):
        """Wrapper method to select XML file before signing"""
        xml_path = filedialog.askopenfilename(
            title="Select XML file to sign",
            filetypes=[("XML files", "*.xml")]
        )
        if xml_path:
            self.sign_xml(xml_path)

    def sign_xml(self, xml_file):
        """Sign XML file using DSA-SHA1 (based on reference implementation)"""
        try:
            # Generate keys if not already generated
            if not self.xmlsec_key:
                self.generate_keys_internal()
            
            # Parse XML
            parser = etree.XMLParser(remove_blank_text=True)
            doc = etree.parse(xml_file, parser)
            
            # Create signing context
            sign_ctx = xmlsec.SignatureContext()
            sign_ctx.key = self.xmlsec_key

            # Create the Signature node
            signature_node = xmlsec.template.create(
                doc,
                xmlsec.constants.TransformExclC14N,  # CanonicalizationMethod
                xmlsec.constants.TransformDsaSha1    # SignatureMethod: DSA-SHA1
            )
            doc.getroot().append(signature_node)

            # Add the Reference node with DigestMethod
            ref = xmlsec.template.add_reference(
                signature_node,
                xmlsec.constants.TransformSha1,  # DigestMethod: SHA1
                uri=""
            )
            xmlsec.template.add_transform(ref, xmlsec.constants.TransformEnveloped)

            # Add KeyInfo with DSAKeyValue
            key_info = xmlsec.template.ensure_key_info(signature_node)
            key_value = etree.SubElement(key_info, "{http://www.w3.org/2000/09/xmldsig#}KeyValue")
            dsa_key_value = etree.SubElement(key_value, "{http://www.w3.org/2000/09/xmldsig#}DSAKeyValue")
            
            # Create P, Q, G, Y elements
            p_elem = etree.SubElement(dsa_key_value, "{http://www.w3.org/2000/09/xmldsig#}P")
            q_elem = etree.SubElement(dsa_key_value, "{http://www.w3.org/2000/09/xmldsig#}Q")
            g_elem = etree.SubElement(dsa_key_value, "{http://www.w3.org/2000/09/xmldsig#}G")
            y_elem = etree.SubElement(dsa_key_value, "{http://www.w3.org/2000/09/xmldsig#}Y")

            # Load public key and get DSA parameters
            public_key = serialization.load_pem_public_key(self.public_key_pem)
            numbers = public_key.public_numbers()
            
            # Encode parameters to base64
            p_elem.text = base64.b64encode(
                numbers.parameter_numbers.p.to_bytes(
                    (numbers.parameter_numbers.p.bit_length() + 7) // 8, 
                    byteorder='big'
                )
            ).decode('utf-8')
            
            q_elem.text = base64.b64encode(
                numbers.parameter_numbers.q.to_bytes(
                    (numbers.parameter_numbers.q.bit_length() + 7) // 8, 
                    byteorder='big'
                )
            ).decode('utf-8')
            
            g_elem.text = base64.b64encode(
                numbers.parameter_numbers.g.to_bytes(
                    (numbers.parameter_numbers.g.bit_length() + 7) // 8, 
                    byteorder='big'
                )
            ).decode('utf-8')
            
            y_elem.text = base64.b64encode(
                numbers.y.to_bytes(
                    (numbers.y.bit_length() + 7) // 8, 
                    byteorder='big'
                )
            ).decode('utf-8')

            # Sign the document
            sign_ctx.sign(signature_node)

            # Save signed XML
            signed_path = xml_file.replace(".xml", "_signed.xml")
            with open(signed_path, 'wb') as f:
                f.write(etree.tostring(doc, xml_declaration=True, encoding='UTF-8'))

            messagebox.showinfo("Signed!", f"DSA-SHA1 signed XML saved:\n{signed_path}")

        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            messagebox.showerror("Signing Failed", f"Error: {str(e)}\n\nDetails:\n{error_details}")

    def verify_xml(self):
        """Verify signed XML"""
        path = filedialog.askopenfilename(
            title="Select signed XML to verify",
            filetypes=[("XML files", "*.xml")]
        )
        if not path:
            return

        try:
            # Generate keys if not already generated
            if not self.xmlsec_key:
                self.generate_keys_internal()
                
            doc = etree.parse(path)
            signature_node = xmlsec.tree.find_node(doc, xmlsec.constants.NodeSignature)
            
            if signature_node is None:
                messagebox.showerror("Error", "No signature found in XML file!")
                return
            
            ctx = xmlsec.SignatureContext()
            key = xmlsec.Key.from_memory(self.public_key_pem, xmlsec.constants.KeyDataFormatPem)
            ctx.key = key
            ctx.verify(signature_node)
            messagebox.showinfo("Valid", "DSA signature is VALID and trusted!")
        except Exception as e:
            messagebox.showerror("Invalid", f"Signature verification failed:\n{str(e)}")

    def save_config(self):
        config = {"elements": []}

        for e in self.elements:
            if not e["frame"].winfo_exists():
                continue

            if e["type"] == "single":
                config["elements"].append({
                    "type": "single",
                    "node": self._safe_get(e["node"]),
                    "row": self._safe_get(e["row"]),
                    "col": self._safe_get(e["col"]),
                    "fixed": self._safe_get(e["fixed"])
                })

            elif e["type"] == "repeated":
                fields = []
                for f in e["fields"]:
                    node = self._safe_get(f["node"])
                    col = self._safe_get(f["col"])
                    off = self._safe_get(f["offset"])
                    if node.strip():
                        fields.append({"node": node, "col": col, "offset": off})
                config["elements"].append({
                    "type": "repeated",
                    "name": self._safe_get(e["name"]),
                    "start": self._safe_get(e["start"]),
                    "end": self._safe_get(e["end"]),
                    "fields": fields
                })

            elif e["type"] == "nested":
                subs = []
                for s in e["subnodes"]:
                    tag = self._safe_get(s["tag"])
                    val = self._safe_get(s["value"])
                    row = self._safe_get(s["row"])
                    col = self._safe_get(s["col"])
                    if tag.strip():
                        subs.append({"tag": tag, "value": val, "row": row, "col": col})
                config["elements"].append({
                    "type": "nested",
                    "block_name": self._safe_get(e["block_name"]),
                    "subnodes": subs
                })

        if not config["elements"]:
            messagebox.showwarning("Empty", "Nothing to save")
            return

        path = filedialog.asksaveasfilename(defaultextension=".json", filetypes=[("JSON Template", "*.json")])
        if path:
            with open(path, "w", encoding="utf-8") as f:
                json.dump(config, f, indent=4)
            messagebox.showinfo("Saved", f"Template saved!\n{path}")

    def load_config(self):
        path = filedialog.askopenfilename(filetypes=[("JSON Template", "*.json")])
        if not path: 
            return

        try:
            with open(path, 'r', encoding='utf-8') as f:
                config = json.load(f)
            
            # Clear existing elements
            for e in self.elements[:]:
                if e["frame"].winfo_exists():
                    e["frame"].destroy()
            self.elements.clear()

            for item in config.get("elements", []):
                t = item["type"]
                if t == "single":
                    self.add_single()
                    el = self.elements[-1]
                    el["node"].insert(0, item.get("node", ""))
                    el["row"].insert(0, item.get("row", ""))
                    el["col"].insert(0, item.get("col", ""))
                    el["fixed"].insert(0, item.get("fixed", ""))

                elif t == "repeated":
                    self.add_repeated()
                    el = self.elements[-1]
                    el["name"].insert(0, item.get("name", ""))
                    el["start"].delete(0, "end")
                    el["start"].insert(0, item.get("start", "2"))
                    el["end"].delete(0, "end")
                    el["end"].insert(0, item.get("end", ""))
                    
                    # Clear default field
                    for f in el["fields"][:]:
                        f["node"].master.destroy()
                    el["fields"].clear()
                    
                    # Add loaded fields
                    for fd in item.get("fields", []):
                        ff = ctk.CTkFrame(el["fields_frame"])
                        ff.pack(fill="x", pady=3)
                        ne = ctk.CTkEntry(ff, width=220)
                        ne.insert(0, fd.get("node", ""))
                        ce = ctk.CTkEntry(ff, width=80)
                        ce.insert(0, fd.get("col", ""))
                        oe = ctk.CTkEntry(ff, width=100)
                        oe.insert(0, fd.get("offset", "0"))
                        ne.pack(side="left", padx=5)
                        ce.pack(side="left", padx=5)
                        oe.pack(side="left", padx=5)
                        
                        field_dict = {"node": ne, "col": ce, "offset": oe}
                        ctk.CTkButton(ff, text="X", width=40, fg_color="red", 
                                    command=lambda f=ff, fd=field_dict: (f.destroy(), el["fields"].remove(fd) if fd in el["fields"] else None)).pack(side="right", padx=5)
                        el["fields"].append(field_dict)

                elif t == "nested":
                    self.add_nested()
                    el = self.elements[-1]
                    el["block_name"].insert(0, item.get("block_name", ""))
                    
                    # Clear default subnode
                    for s in el["subnodes"][:]:
                        s["tag"].master.destroy()
                    el["subnodes"].clear()
                    
                    # Add loaded subnodes
                    for sub in item.get("subnodes", []):
                        ff = ctk.CTkFrame(el["subs_frame"])
                        ff.pack(fill="x", pady=3)
                        te = ctk.CTkEntry(ff, width=250)
                        te.insert(0, sub.get("tag", ""))
                        ve = ctk.CTkEntry(ff, width=200)
                        ve.insert(0, sub.get("value", ""))
                        re = ctk.CTkEntry(ff, width=80)
                        re.insert(0, sub.get("row", ""))
                        ce = ctk.CTkEntry(ff, width=80)
                        ce.insert(0, sub.get("col", ""))
                        te.pack(side="left", padx=5)
                        ve.pack(side="left", padx=5)
                        re.pack(side="left", padx=5)
                        ce.pack(side="left", padx=5)
                        
                        sub_dict = {"tag": te, "value": ve, "row": re, "col": ce}
                        ctk.CTkButton(ff, text="X", width=40, fg_color="red",
                                    command=lambda f=ff, sd=sub_dict: (f.destroy(), el["subnodes"].remove(sd) if sd in el["subnodes"] else None)).pack(side="right", padx=5)
                        el["subnodes"].append(sub_dict)

            messagebox.showinfo("Loaded", "Template loaded successfully!")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to load template:\n{str(e)}")


if __name__ == "__main__":
    app = ExcelToXMLUltimate()
    app.mainloop()